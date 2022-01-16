# anade region a los edges
import openpyxl
from pathlib import Path


def main():
    loc = 'no_repeated_rows.xlsx'
    wb_obj = openpyxl.load_workbook(loc)
    print(wb_obj.sheetnames)
    sheet_obj = wb_obj.active


    loc1 = 'NE Report_2021-12-01_17-31-54_1 sdh.xlsx'
    wb_obj1 = openpyxl.load_workbook(loc1)
    sheet_obj1 = wb_obj1.active

    # Will print a particular row value
    for j in range(9, sheet_obj.max_row + 1):
        cell = sheet_obj.cell(row=j, column=6)
        cell_to_update = sheet_obj.cell(row=j, column=23)
        net = search_net(sheet_obj1, cell.value)
        if net:
            print(cell.value + '->' + net)
            cell_to_update.value = net
        else:
            cell = sheet_obj.cell(row=j, column=8)
            net = search_net(sheet_obj1, cell.value)
            print(cell.value + '->' + net)
            cell_to_update.value = net

    wb_obj.save('no_repeated_rows_with_region.xlsx')


def search_net(sheet_obj, search_string):
    # Will print a particular row value
    for j in range(5, sheet_obj.max_row + 1):
        if sheet_obj.cell(row=j, column=1).value == search_string:
            return sheet_obj.cell(row=j, column=8).value

    return None


if __name__ == '__main__':
    main()
