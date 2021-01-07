# This is a sample Python script.
import xlrd
import openpyxl
from pathlib import Path
import networkx as nx
import matplotlib.pyplot as plt


def main():
    G = nx.Graph()
    xlsx_file = Path('SimData', 'play_data.xlsx')
    loc = 'REGION 1 TOPOLOGIA.xlsx'
    wb_obj = openpyxl.load_workbook(loc)
    sheet_obj = wb_obj.active

    max_col = sheet_obj.max_column

    # Will print a particular row value
    for j in range(9, 479):
        row = ''
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=j, column=i)
            row = row + cell_obj.value

            if i == 3:
                print(cell_obj.value)
                if cell_obj.value:
                    G.add_node(cell_obj.value)
                    source = cell_obj.value

            if i == 6:
                print(cell_obj.value)
                if cell_obj.value:
                    G.add_node(cell_obj.value)
                    target = cell_obj.value

        if source and target and not G.has_edge(source, target) and not G.has_edge(target, source):
            G.add_edge(source, target)

        # print(row)
    plt.figure(3, figsize=(50, 50))
    nx.draw_networkx(G, arrowsize=6, with_labels=True, node_size=10, font_size=5)
    plt.savefig("path_graph1.pdf")
    plt.show()


if __name__ == '__main__':
    main()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
