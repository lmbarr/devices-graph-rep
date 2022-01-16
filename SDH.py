import openpyxl
from pathlib import Path
import networkx as nx
import matplotlib.pyplot as plt


def process(sheet_obj, region):
    G = nx.Graph()

    # Will print a particular row value

    for j in range(9, sheet_obj.max_row + 1):
        source = sheet_obj.cell(row=j, column=6)
        target = sheet_obj.cell(row=j, column=8)
        region_read = sheet_obj.cell(row=j, column=23)
        print(region, region_read)
        if region == region_read.value:

            if source.value:
                G.add_node(source.value)

            if target.value:
                G.add_node(target.value)

            if source.value and target.value and \
                    source.value != target.value and \
                    not G.has_edge(source.value, target.value) and \
                    not G.has_edge(target.value, source.value):
                G.add_edge(source.value, target.value)

    plt.figure(3, figsize=(50, 50))
    nx.draw_networkx(G, arrowsize=6, with_labels=True, node_size=10, font_size=4)
    plt.savefig(region + ".pdf")
    plt.clf()


def main():
    loc = 'no_repeated_rows_with_region.xlsx'
    wb_obj = openpyxl.load_workbook(loc)

    for i in range(1, 8, 1):
        process(wb_obj.active, 'REGION ' + str(i))


if __name__ == '__main__':
    main()

