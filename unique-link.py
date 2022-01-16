import openpyxl
from pathlib import Path
import networkx as nx
import matplotlib.pyplot as plt


def main():
    G = nx.Graph()
    loc = 'Cable_Connection_Report_2021-12-01_17-28-49 sdh.xlsx'
    wb_obj = openpyxl.load_workbook(loc)
    sheet_obj = wb_obj.active

    max_col = sheet_obj.max_column

    # Will print a particular row value
    for j in range(9, 1797):
        row = ''
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=j, column=i)
            row = row + cell_obj.value

            if i == 6:
                if cell_obj.value:
                    G.add_node(cell_obj.value)
                    source = cell_obj.value
                    print(source)

            if i == 8:
                if cell_obj.value:
                    G.add_node(cell_obj.value)
                    target = cell_obj.value
                    print(target)


        if source and target and not G.has_edge(source, target) and not G.has_edge(target, source):
            G.add_edge(source, target)

   
    with open('grafo.txt', 'a') as out:
        for i in G.edges:
            print(i)
            out.write(','.join(i)+ '\n')
"""         # print(row)
    plt.figure(3, figsize=(50, 50))
    nx.draw_networkx(G, arrowsize=6, with_labels=True, node_size=10, font_size=5)
    plt.savefig(sheet_name + ".pdf")
    plt.clf() """


if __name__ == '__main__':
    main()

