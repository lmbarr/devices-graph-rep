# This is a sample Python script.
import openpyxl
from pathlib import Path


def main():
    loc = 'clasificación de redes.xlsx'
    wb_obj = openpyxl.load_workbook(loc)
    print(wb_obj.sheetnames)
    sheet_obj = wb_obj['Hoja7']

    # Will print a particular row value
    for j in range(9, 23213):
        cell = sheet_obj.cell(row=j, column=3)
        cell_to_update = sheet_obj.cell(row=j, column=1)
        net = search_net(wb_obj, cell.value)
        if net:
            print(cell.value + '->' + net)
            cell_to_update.value = net
        else:
            cell = sheet_obj.cell(row=j, column=2)
            net = search_net(wb_obj, cell.value)
            print(cell.value + '->' + net)
            cell_to_update.value = net

    wb_obj.save('clasificación de redes updated.xlsx')


def search_net(wb, search_string):
    sheet = wb['CLASIFICACIÓN']

    # Will print a particular row value
    for j in range(5, 423):
        if sheet.cell(row=j, column=1).value == search_string:
            return sheet.cell(row=j, column=2).value

    return None


if __name__ == '__main__':
    main()
