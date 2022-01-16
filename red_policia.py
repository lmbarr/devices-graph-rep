import openpyxl
import networkx as nx
import matplotlib.pyplot as plt


def process(sheet_obj, subnet):
    G = nx.Graph()

    # Will print a particular row value

    for j in range(2, sheet_obj.max_row + 1):
        source = sheet_obj.cell(row=j, column=2)
        target = sheet_obj.cell(row=j, column=3)
        subnet_read = sheet_obj.cell(row=j, column=1)

        if subnet == subnet_read.value:

            if source.value:
                G.add_node(source.value)

            if target.value:
                G.add_node(target.value)

            if source.value and target.value and \
                    source.value != target.value and \
                    not G.has_edge(source.value, target.value) and \
                    not G.has_edge(target.value, source.value):
                G.add_edge(source.value, target.value)

        # print(row)
    plt.figure(3, figsize=(30, 30))
    nx.draw_networkx(G, arrowsize=6, with_labels=True, node_size=10, font_size=10)
    plt.savefig(subnet + ".pdf")
    plt.clf()


def main():
    loc = 'ENLACES MW SIAE.xlsx'
    wb_obj = openpyxl.load_workbook(loc)

    subnet_list = set()
    for cell in wb_obj.active['A']:
        if cell.value and cell.value != '':
            subnet_list.add(cell.value)
    
    print(subnet_list)

    for subnet in subnet_list:
        process(wb_obj.active, subnet)


if __name__ == '__main__':
    main()

