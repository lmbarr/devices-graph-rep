import openpyxl
import networkx as nx
import matplotlib.pyplot as plt


def process(sheet_obj, subnet):
    G = nx.Graph()

    # Will print a particular row value

    for j in range(9, sheet_obj.max_row + 1):
        source = sheet_obj.cell(row=j, column=6)
        target = sheet_obj.cell(row=j, column=8)
        subnet_read = sheet_obj.cell(row=j, column=23)

        if subnet == subnet_read.value:

            if source.value:
                G.add_node(source.value)

            if target.value:
                G.add_node(target.value)

            if source.value and target.value and \
                    source.value != target.value and \
                    not G.has_edge(source.value, target.value) and \
                    not G.has_edge(target.value, source.value):
                G.add_edge(source.value, target.value)

        # print(row)
    plt.figure(3, figsize=(30, 30))
    nx.draw_networkx(G, arrowsize=6, with_labels=True, node_size=10, font_size=10)
    plt.savefig(subnet + ".pdf")
    plt.clf()


def main():
    loc = 'Cable_Connection_Report_2021-12-01_17-42-09 dwdm_only_wdm_no_repeated_rows_with_subnet.xlsx'
    wb_obj = openpyxl.load_workbook(loc)

    subnet_list = set()
    for cell in wb_obj.active['W']:
        if cell.value and cell.value != '':
            subnet_list.add(cell.value)
    
    print(subnet_list)

    for subnet in subnet_list:
        process(wb_obj.active, subnet)


if __name__ == '__main__':
    main()

