import openpyxl
from pathlib import Path
file_name_ = 'Cable_Connection_Report_2021-12-01_17-42-09 dwdm'

def main():
    file_name = file_name_ + '_only_wdm'
    wb_obj = openpyxl.load_workbook(file_name + '.xlsx')
    sheet_obj = wb_obj.active

    for j in range(sheet_obj.max_row, 8, -1):            
        #print(j)
        cell_obj_source = sheet_obj.cell(row=j, column=6)
        cell_obj_target = sheet_obj.cell(row=j, column=8)

        source = cell_obj_source.value
        #print(source)

        target = cell_obj_target.value
        print(j, target, source)

        for i in range(j - 1, 8, -1):
            cell_obj_source_aux = sheet_obj.cell(row=i, column=6)
            cell_obj_target_aux = sheet_obj.cell(row=i, column=8)

            source_aux = cell_obj_source_aux.value
            #print(source_aux)

            target_aux = cell_obj_target_aux.value
            print(i, source_aux, target_aux)
            #print(source , target , source_aux , target_aux)
            if source and target and source_aux and target_aux and (
                (source == source_aux and target == target_aux) or 
                (source == target_aux and target == source_aux)):
                print('hay repetidos')
                sheet_obj.delete_rows(i, 1)
                
    # borrar self edges
    for j in range(sheet_obj.max_row, 8, -1):
        cell_obj_source = sheet_obj.cell(row=j, column=6)
        cell_obj_target = sheet_obj.cell(row=j, column=8)
        source = cell_obj_source.value
        target = cell_obj_target.value

        if source == target:
            sheet_obj.delete_rows(j, 1)

    wb_obj.save(file_name + '_no_repeated_rows.xlsx')

def deleteRowsNotDWM():
    wb_obj = openpyxl.load_workbook(file_name_ + '.xlsx')
    sheet_obj = wb_obj.active

    for j in range(sheet_obj.max_row, 8, -1):            
        cell = sheet_obj.cell(row=j, column=2)
        if cell.value != 'WDM' and cell.value != 'WDM CORD':
            sheet_obj.delete_rows(j, 1)

    wb_obj.save(file_name_ + '_only_wdm.xlsx')


if __name__ == '__main__':
    deleteRowsNotDWM()
    main()

